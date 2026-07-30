"""
Excel 报表导出模块
使用 openpyxl 生成带格式的质检报表。
"""

import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

logger = logging.getLogger("VideoQC.Exporter")


class ExcelExporter:
    """Excel 报表导出器"""

    # 颜色定义
    COLOR_PASS = "34A853"
    COLOR_WARNING = "FBBC04"
    COLOR_ERROR = "EA4335"
    COLOR_HEADER_BG = "F1F3F4"
    COLOR_HEADER_FONT = "1A1A2E"
    COLOR_ALT_ROW = "F8F9FA"

    def __init__(self):
        pass

    def export(self, results, consistency, output_path):
        """
        导出质检报告到 Excel

        Args:
            results: 检测结果列表
            consistency: 一致性校验结果
            output_path: 输出文件路径
        """
        logger.info(f"开始导出 Excel: {output_path}")

        wb = Workbook()

        # Sheet 1: 质检汇总
        self._create_summary_sheet(wb, results, consistency)

        # Sheet 2: 详细检测结果
        self._create_detail_sheet(wb, results)

        # Sheet 3: 异常详情
        self._create_anomaly_sheet(wb, results)

        # 删除默认空白 sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        wb.save(output_path)
        logger.info(f"Excel 导出完成: {output_path}")

    def _create_summary_sheet(self, wb, results, consistency):
        """创建汇总表"""
        ws = wb.active
        ws.title = "质检汇总"

        # 样式定义
        header_font = Font(name="微软雅黑", size=11, bold=True, color=self.COLOR_HEADER_FONT)
        header_fill = PatternFill(start_color=self.COLOR_HEADER_BG, end_color=self.COLOR_HEADER_BG, fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin", color="DADCE0"),
            right=Side(style="thin", color="DADCE0"),
            top=Side(style="thin", color="DADCE0"),
            bottom=Side(style="thin", color="DADCE0"),
        )

        pass_fill = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
        warn_fill = PatternFill(start_color="FEF7E0", end_color="FEF7E0", fill_type="solid")
        error_fill = PatternFill(start_color="FCE8E6", end_color="FCE8E6", fill_type="solid")

        # 标题行
        ws.merge_cells("A1:O1")
        title_cell = ws["A1"]
        title_cell.value = f"VideoQC Pro 视频质检报告 - {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        title_cell.font = Font(name="微软雅黑", size=16, bold=True, color="1A73E8")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 40

        # 表头（第3行）
        headers = [
            "序号", "文件名", "文件大小(MB)", "时长(s)",
            "视频编码", "分辨率", "帧率", "视频码率(kbps)",
            "音频编码", "采样率(Hz)", "声道", "音频码率(kbps)",
            "黑帧", "夹帧/跳帧", "黑边", "静音", "综合判定"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # 冻结表头
        ws.freeze_panes = "A4"

        # 数据行
        for i, result in enumerate(results):
            row = i + 4
            metadata = result.get("metadata") or {}
            video = metadata.get("video") or {}
            audio = metadata.get("audio") or {}

            # 检测结果汇总
            bf_ok = not (result.get("black_frame") or {}).get("has_black_frames", False)
            ff_ok = not (result.get("flash_frame") or {}).get("has_flash_frames", False)
            bb_ok = not (result.get("black_border") or {}).get("has_black_border", False)
            sd_ok = not (result.get("silence") or {}).get("has_silence", False)

            def status_text(ok):
                return "✓ 通过" if ok else "✗ 异常"

            row_data = [
                i + 1,
                result.get("filename", ""),
                metadata.get("filesize_mb", ""),
                metadata.get("duration", ""),
                video.get("codec", ""),
                video.get("aspect_ratio", ""),
                video.get("fps", ""),
                round(video.get("bitrate", 0) / 1000, 1) if video.get("bitrate") else "",
                audio.get("codec", ""),
                audio.get("sample_rate", ""),
                audio.get("channels", ""),
                round(audio.get("bitrate", 0) / 1000, 1) if audio.get("bitrate") else "",
                status_text(bf_ok),
                status_text(ff_ok),
                status_text(bb_ok),
                status_text(sd_ok),
                self._get_status_label(result.get("overall_status", "pending")),
            ]

            alt_fill = PatternFill(start_color=self.COLOR_ALT_ROW, end_color=self.COLOR_ALT_ROW, fill_type="solid") if i % 2 == 0 else None

            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.alignment = cell_align
                cell.border = thin_border
                if alt_fill:
                    cell.fill = alt_fill

            # 综合判定着色
            overall_cell = ws.cell(row=row, column=17)
            status = result.get("overall_status", "pending")
            if status == "pass":
                overall_cell.fill = pass_fill
                overall_cell.font = Font(name="微软雅黑", color="137333", bold=True)
            elif status in ("warning",):
                overall_cell.fill = warn_fill
                overall_cell.font = Font(name="微软雅黑", color="B06000", bold=True)
            elif status in ("fail", "error"):
                overall_cell.fill = error_fill
                overall_cell.font = Font(name="微软雅黑", color="C5221F", bold=True)

        # 列宽调整
        col_widths = [6, 30, 10, 8, 12, 14, 8, 12, 12, 10, 6, 12, 10, 10, 10, 10, 12]
        for col, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # 统计摘要
        stats_row = len(results) + 5
        pass_count = sum(1 for r in results if r.get("overall_status") == "pass")
        warn_count = sum(1 for r in results if r.get("overall_status") == "warning")
        fail_count = sum(1 for r in results if r.get("overall_status") in ("fail", "error"))

        ws.merge_cells(f"A{stats_row}:C{stats_row}")
        ws.cell(row=stats_row, column=1, value="统计:").font = Font(bold=True, size=12)
        ws.cell(row=stats_row, column=4, value=f"总计: {len(results)}").font = Font(bold=True)
        ws.cell(row=stats_row, column=5, value=f"通过: {pass_count}").font = Font(color="137333", bold=True)
        ws.cell(row=stats_row, column=6, value=f"警告: {warn_count}").font = Font(color="B06000", bold=True)
        ws.cell(row=stats_row, column=7, value=f"不合格: {fail_count}").font = Font(color="C5221F", bold=True)

    def _create_detail_sheet(self, wb, results):
        """创建详细检测结果表"""
        ws = wb.create_sheet("检测详情")

        header_font = Font(name="微软雅黑", size=11, bold=True)
        header_fill = PatternFill(start_color=self.COLOR_HEADER_BG, end_color=self.COLOR_HEADER_BG, fill_type="solid")

        headers = [
            "文件名", "检测项", "检测结果", "异常详情", "严重程度"
        ]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill

        ws.freeze_panes = "A2"

        row = 2
        for result in results:
            filename = result.get("filename", "")

            # 黑帧
            bf = result.get("black_frame", {})
            if bf:
                segs = bf.get("segments", [])
                detail = "; ".join([f"{s.get('start_time', 0)}s-{s.get('end_time', 0)}s ({s.get('severity', '')})" for s in segs[:5]]) if segs else "无"
                ws.cell(row=row, column=1, value=filename)
                ws.cell(row=row, column=2, value="黑帧检测")
                ws.cell(row=row, column=3, value="异常" if segs else "通过")
                ws.cell(row=row, column=4, value=detail)
                ws.cell(row=row, column=5, value="错误" if any(s.get("severity") == "错误" for s in segs) else ("警告" if segs else "通过"))
                row += 1

            # 夹帧/跳帧
            ff = result.get("flash_frame", {})
            if ff:
                cands = ff.get("candidates", [])
                detail = "; ".join([
                    f"{c.get('type', '夹帧')}帧{c.get('start_frame', 0)}~{c.get('end_frame', 0)}"
                    f"({c.get('span_frames', 1)}帧/{c.get('duration_ms', 0):.0f}ms)"
                    for c in cands[:5]
                ]) if cands else "无"
                ws.cell(row=row, column=1, value=filename)
                ws.cell(row=row, column=2, value="夹帧/跳帧检测")
                ws.cell(row=row, column=3, value="异常" if cands else "通过")
                ws.cell(row=row, column=4, value=detail)
                ws.cell(row=row, column=5, value="高危 人工复核" if cands else "通过")
                row += 1

            # 黑边
            bb = result.get("black_border", {})
            if bb:
                ws.cell(row=row, column=1, value=filename)
                ws.cell(row=row, column=2, value="黑边检测")
                ws.cell(row=row, column=3, value="异常" if bb.get("has_black_border") else "通过")
                ws.cell(row=row, column=4, value=bb.get("border_type", "") + f" (有效占比: {bb.get('avg_valid_ratio', 0)})")
                ws.cell(row=row, column=5, value="警告" if bb.get("has_black_border") else "通过")
                row += 1

            # 静音
            sd = result.get("silence", {})
            if sd:
                segs = sd.get("segments", [])
                detail = "; ".join([f"{s.get('start', 0)}s-{s.get('end', 0)}s ({s.get('severity', '')})" for s in segs[:5]]) if segs else "无"
                ws.cell(row=row, column=1, value=filename)
                ws.cell(row=row, column=2, value="静音检测")
                ws.cell(row=row, column=3, value="异常" if segs else "通过")
                ws.cell(row=row, column=4, value=detail)
                ws.cell(row=row, column=5, value="错误" if any(s.get("severity") == "错误" for s in segs) else ("警告" if segs else "通过"))
                row += 1

        # 列宽
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 60
        ws.column_dimensions["E"].width = 12

    def _create_anomaly_sheet(self, wb, results):
        """创建异常详情表"""
        ws = wb.create_sheet("异常时间点")

        header_font = Font(name="微软雅黑", size=11, bold=True)
        header_fill = PatternFill(start_color=self.COLOR_HEADER_BG, end_color=self.COLOR_HEADER_BG, fill_type="solid")

        headers = ["文件名", "异常类型", "开始时间", "结束时间", "时长(s)", "严重程度"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill

        ws.freeze_panes = "A2"

        row = 2
        for result in results:
            filename = result.get("filename", "")

            # 黑帧时间点
            bf = result.get("black_frame", {})
            for seg in bf.get("segments", []):
                ws.cell(row=row, column=1, value=filename)
                ws.cell(row=row, column=2, value="黑帧")
                ws.cell(row=row, column=3, value=seg.get("start_time", ""))
                ws.cell(row=row, column=4, value=seg.get("end_time", ""))
                ws.cell(row=row, column=5, value=seg.get("duration", ""))
                ws.cell(row=row, column=6, value=seg.get("severity", ""))
                row += 1

            # 夹帧/跳帧时间点
            ff = result.get("flash_frame", {})
            for cand in ff.get("candidates", []):
                ws.cell(row=row, column=1, value=filename)
                ws.cell(row=row, column=2, value=f"{cand.get('type', '夹帧')}(帧{cand.get('start_frame', '')}~{cand.get('end_frame', '')})")
                ws.cell(row=row, column=3, value=cand.get("start_time", ""))
                ws.cell(row=row, column=4, value=cand.get("end_time", ""))
                ws.cell(row=row, column=5, value=cand.get("span_frames", 1))
                ws.cell(row=row, column=6, value=f"{cand.get('confidence_level', '')} {cand.get('confidence', '?')}%")
                row += 1

            # 静音时间点
            sd = result.get("silence", {})
            for seg in sd.get("segments", []):
                ws.cell(row=row, column=1, value=filename)
                ws.cell(row=row, column=2, value="静音")
                ws.cell(row=row, column=3, value=seg.get("start", ""))
                ws.cell(row=row, column=4, value=seg.get("end", ""))
                ws.cell(row=row, column=5, value=seg.get("duration", ""))
                ws.cell(row=row, column=6, value=seg.get("severity", ""))
                row += 1

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 10
        ws.column_dimensions["F"].width = 10

    def _get_status_label(self, status):
        labels = {
            "pass": "🟢 合格",
            "warning": "🟡 警告",
            "fail": "🔴 不合格",
            "error": "🔴 错误",
            "pending": "⚪ 待检",
            "cancelled": "⏹ 已取消",
        }
        return labels.get(status, status)
