"""
多版本对比弹窗
支持拖放文件夹、展示横向对比结果、点击展开查看各版本详细参数、导出 Excel。

表格设计：
- 每个视频编号一行，各检测项只显示 ✓/✗（固定窄列宽）
- 点击展开显示各版本的详细参数值和不一致详情
- 支持拖放文件夹、自然排序、仅显示不一致行过滤
"""
import os
import logging
from datetime import datetime
from typing import List

from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QTreeWidget, QTreeWidgetItem, QHeaderView, QFileDialog,
    QMessageBox, QProgressBar, QListWidget, QListWidgetItem,
    QAbstractItemView, QWidget, QApplication, QCheckBox,
)
from PySide6.QtCore import Qt, QThread, Signal
from PySide6.QtGui import QDragEnterEvent, QDropEvent, QColor

from core.multi_version_compare import (
    MultiVersionComparator, CompareResult, GroupResult, VersionFile, natural_sort_key,
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger("VideoQC.MultiVersionCompareDialog")


# ── 列定义：每个检测项 ──
COMPARE_COLUMNS = [
    ("时长", "duration",   lambda v: f"{v.duration:.2f}s" if v.duration else "-"),
    ("帧率", "fps",        lambda v: f"{v.fps:.2f}" if v.fps else "-"),
    ("编码", "codec",      lambda v: v.codec or "-"),
    ("帧数", "frame_count", lambda v: f"{v.frame_count:,}" if v.frame_count is not None else "-"),
    ("格式", "format",     lambda v: v.format_name or "-"),
    ("分辨率", "resolution", lambda v: f"{v.width}×{v.height}" if v.width and v.height else "-"),
]

# 列总数：编号(1) + 检测项(6) + 状态(1) = 8
COL_COUNT = 1 + len(COMPARE_COLUMNS) + 1


class CompareThread(QThread):
    """后台对比线程"""
    progress = Signal(int, str)
    finished = Signal(object)
    error = Signal(str)

    def __init__(self, folders, parent=None):
        super().__init__(parent)
        self.folders = folders

    def run(self):
        try:
            comparator = MultiVersionComparator(progress_callback=self._on_progress)
            result = comparator.compare(self.folders)
            self.finished.emit(result)
        except Exception as e:
            logger.exception("多版本对比失败")
            self.error.emit(str(e))

    def _on_progress(self, percent, filename):
        self.progress.emit(percent, filename)


class MultiVersionCompareDialog(QDialog):
    """多版本视频对比对话框

    设计要点：
    - 每个视频编号一行，各检测项只显示 ✓/✗ 一致性标记（固定窄列宽）
    - 点击行展开，子行显示各版本的具体参数值和不一致详情
    - 支持拖放文件夹、自然排序、仅显示不一致行过滤。
    """

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("多版本视频对比")
        self.setMinimumSize(1180, 720)
        self.setModal(True)
        self._result: CompareResult = None
        self._thread: CompareThread = None
        self._show_inconsistent_only = False

        # 启用整窗拖放接收
        self.setAcceptDrops(True)

        self._init_ui()

    # ── UI 构建 ──────────────────────────────────────────────
    def _init_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(10)
        layout.setContentsMargins(16, 16, 16, 16)

        # ── 使用说明 ──
        help_label = QLabel(
            "使用说明：每个文件夹代表一个版本（如不同音频/水印），各版本内文件名相同的视频会被横向对比时长/帧率/编码/帧数等。"
            "点击行可展开查看各版本具体参数值和不一致详情。"
        )
        help_label.setWordWrap(True)
        help_label.setProperty("cssClass", "text-secondary-xs")
        layout.addWidget(help_label)

        # ── 文件夹列表 + 操作按钮 ──
        top = QHBoxLayout()
        top.setSpacing(8)

        list_container = QWidget()
        list_container.setProperty("cssClass", "card")
        list_container.setAcceptDrops(True)
        list_container.dragEnterEvent = self._list_drag_enter
        list_container.dragMoveEvent = self._list_drag_enter
        list_container.dropEvent = self._list_drop
        list_layout = QVBoxLayout(list_container)
        list_layout.setContentsMargins(12, 8, 12, 10)
        list_layout.setSpacing(6)

        self.folder_list = QListWidget()
        self.folder_list.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        self.folder_list.setMaximumHeight(110)
        self.folder_list.setAlternatingRowColors(True)
        list_layout.addWidget(self.folder_list)

        self._drop_hint = QLabel("拖入文件夹 或 点击右侧「添加文件夹」按钮")
        self._drop_hint.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._drop_hint.setProperty("cssClass", "text-secondary-xs")
        self._drop_hint.setStyleSheet("padding: 18px;")
        list_layout.addWidget(self._drop_hint)
        self._drop_hint.setVisible(True)

        top.addWidget(list_container, 1)

        btns = QVBoxLayout()
        btns.setSpacing(6)
        self.btn_add_folder = QPushButton("添加文件夹")
        self.btn_add_folder.setProperty("cssClass", "primary")
        self.btn_add_folder.clicked.connect(self._add_folder)
        btns.addWidget(self.btn_add_folder)

        self.btn_remove_folder = QPushButton("移除选中")
        self.btn_remove_folder.clicked.connect(self._remove_selected)
        btns.addWidget(self.btn_remove_folder)

        self.btn_clear_folders = QPushButton("清空")
        self.btn_clear_folders.clicked.connect(self._clear_folders)
        btns.addWidget(self.btn_clear_folders)
        btns.addStretch()
        top.addLayout(btns)
        layout.addLayout(top)

        # ── 进度条 + 状态 ──
        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        self.progress_bar.setTextVisible(True)
        self.progress_bar.setFormat("%p%")
        layout.addWidget(self.progress_bar)

        self.status_label = QLabel("请选择 2 个及以上版本文件夹")
        self.status_label.setProperty("cssClass", "text-secondary")
        layout.addWidget(self.status_label)

        # ── 过滤开关 + 摘要 ──
        filter_row = QHBoxLayout()
        self.chk_inconsistent_only = QCheckBox("仅显示不一致行")
        self.chk_inconsistent_only.stateChanged.connect(self._on_filter_changed)
        filter_row.addWidget(self.chk_inconsistent_only)

        self.summary_label = QLabel("")
        self.summary_label.setProperty("cssClass", "text-secondary")
        filter_row.addStretch()
        filter_row.addWidget(self.summary_label)
        layout.addLayout(filter_row)

        # ── 结果树形表格 ──
        self.result_tree = QTreeWidget()
        headers = ["视频编号"] + [c[0] for c in COMPARE_COLUMNS] + ["状态"]
        self.result_tree.setHeaderLabels(headers)
        self.result_tree.setAlternatingRowColors(True)
        self.result_tree.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.result_tree.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.result_tree.setMouseTracking(True)
        self.result_tree.setIndentation(20)
        # 默认全部展开
        self.result_tree.setExpandsOnDoubleClick(True)
        layout.addWidget(self.result_tree, 1)

        # 固定列宽
        tree_header = self.result_tree.header()
        tree_header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)  # 编号
        for c in range(1, 1 + len(COMPARE_COLUMNS)):
            tree_header.setSectionResizeMode(c, QHeaderView.ResizeMode.Fixed)
            tree_header.resizeSection(c, 70)  # ✓/✗ 列固定 70px
        tree_header.setSectionResizeMode(COL_COUNT - 1, QHeaderView.ResizeMode.Stretch)  # 状态列拉满

        # ── 底部按钮 ──
        bottom = QHBoxLayout()
        self.btn_start = QPushButton("开始对比")
        self.btn_start.setProperty("cssClass", "primary")
        self.btn_start.clicked.connect(self._start_compare)
        bottom.addWidget(self.btn_start)

        self.btn_export = QPushButton("导出 Excel")
        self.btn_export.clicked.connect(self._export_excel)
        self.btn_export.setEnabled(False)
        bottom.addWidget(self.btn_export)

        bottom.addStretch()

        self.btn_close = QPushButton("关闭")
        self.btn_close.clicked.connect(self.reject)
        bottom.addWidget(self.btn_close)
        layout.addLayout(bottom)

        self._refresh_drop_hint()

    # ── 拖放支持 ─────────────────────────────────────────────
    def dragEnterEvent(self, event: QDragEnterEvent):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
        else:
            event.ignore()

    def dragMoveEvent(self, event: QDragEnterEvent):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
        else:
            event.ignore()

    def dropEvent(self, event: QDropEvent):
        paths = self._extract_folders(event.mimeData().urls())
        if paths:
            self._add_folder_paths(paths)
            event.acceptProposedAction()
        else:
            event.ignore()

    def _list_drag_enter(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
        else:
            event.ignore()

    def _list_drop(self, event):
        paths = self._extract_folders(event.mimeData().urls())
        if paths:
            self._add_folder_paths(paths)
            event.acceptProposedAction()
        else:
            event.ignore()

    @staticmethod
    def _extract_folders(urls) -> List[str]:
        """从拖放 URL 列表中提取文件夹路径（自动跳过文件）"""
        folders = []
        for url in urls:
            local = url.toLocalFile()
            if not local:
                continue
            if os.path.isdir(local):
                folders.append(os.path.normpath(local))
            elif os.path.isfile(local):
                parent = os.path.dirname(local)
                if parent and os.path.isdir(parent):
                    folders.append(os.path.normpath(parent))
        seen = set()
        unique = []
        for f in folders:
            if f not in seen:
                seen.add(f)
                unique.append(f)
        return unique

    # ── 文件夹管理 ──────────────────────────────────────────
    def _add_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "选择版本文件夹")
        if folder:
            self._add_folder_paths([os.path.normpath(folder)])

    def _add_folder_paths(self, folders: List[str]):
        existing = {
            self.folder_list.item(i).data(Qt.ItemDataRole.UserRole)
            for i in range(self.folder_list.count())
        }
        added = 0
        for folder in folders:
            if folder in existing:
                continue
            item = QListWidgetItem(os.path.basename(folder) or folder)
            item.setData(Qt.ItemDataRole.UserRole, folder)
            item.setToolTip(folder)
            self.folder_list.addItem(item)
            existing.add(folder)
            added += 1
        if added:
            self.status_label.setText(f"已添加 {added} 个文件夹（共 {self.folder_list.count()} 个）")
        self._refresh_drop_hint()

    def _remove_selected(self):
        for item in self.folder_list.selectedItems():
            self.folder_list.takeItem(self.folder_list.row(item))
        self._refresh_drop_hint()

    def _clear_folders(self):
        self.folder_list.clear()
        self._refresh_drop_hint()

    def _refresh_drop_hint(self):
        """文件夹列表为空时显示拖放提示，有内容时隐藏"""
        has_items = self.folder_list.count() > 0
        self._drop_hint.setVisible(not has_items)
        self.folder_list.setMaximumHeight(110 if has_items else 0)

    # ── 对比流程 ─────────────────────────────────────────────
    def _start_compare(self):
        folders = [
            self.folder_list.item(i).data(Qt.ItemDataRole.UserRole)
            for i in range(self.folder_list.count())
        ]
        if len(folders) < 2:
            QMessageBox.warning(self, "提示", "请至少添加 2 个版本文件夹进行对比。")
            return

        self.btn_start.setEnabled(False)
        self.btn_export.setEnabled(False)
        self.progress_bar.setValue(0)
        self.status_label.setText("正在读取元数据并对比...")
        self.result_tree.clear()
        self.summary_label.setText("")

        self._thread = CompareThread(folders, self)
        self._thread.progress.connect(self._on_progress)
        self._thread.finished.connect(self._on_finished)
        self._thread.error.connect(self._on_error)
        self._thread.start()

    def _on_progress(self, percent, filename):
        self.progress_bar.setValue(percent)
        self.status_label.setText(f"正在探测：{filename}  ({percent}%)")

    def _on_finished(self, result: CompareResult):
        self._result = result
        self.progress_bar.setValue(100)
        self.status_label.setText("对比完成")
        self.btn_start.setEnabled(True)
        self.btn_export.setEnabled(True)
        self._populate_tree(result)
        self._update_summary(result)

    def _on_error(self, msg):
        self.btn_start.setEnabled(True)
        self.progress_bar.setValue(0)
        self.status_label.setText(f"对比失败：{msg}")
        QMessageBox.critical(self, "错误", f"对比过程出错：\n{msg}")

    def _on_filter_changed(self, state):
        self._show_inconsistent_only = bool(state)
        if self._result:
            self._populate_tree(self._result)
            self._update_summary(self._result)

    def _update_summary(self, result: CompareResult):
        self.summary_label.setText(
            f"共 {result.total_files} 个文件，"
            f"✓ {result.consistent_groups} 组一致，"
            f"✗ {result.inconsistent_groups} 组不一致"
        )

    # ── 树形表格填充 ────────────────────────────────────────
    def _populate_tree(self, result: CompareResult):
        """填充树形表格：顶层行显示 ✓/✗，子行显示各版本详情"""
        self.result_tree.clear()

        sorted_groups = sorted(result.groups, key=lambda g: natural_sort_key(g.file_id))

        # 过滤不一致行
        if self._show_inconsistent_only:
            sorted_groups = [g for g in sorted_groups if not g.all_consistent]

        for group in sorted_groups:
            top_item = self._create_top_item(group, result.version_names)
            self.result_tree.addTopLevelItem(top_item)

            # 子行：各版本详细参数
            for v in group.versions:
                child = self._create_version_child(v, group, result.version_names)
                top_item.addChild(child)

            # 不一致时默认展开
            if not group.all_consistent:
                top_item.setExpanded(True)

        # 一致组全部展开（方便查看），不一致组已展开
        # 全部展开便于浏览
        # self.result_tree.expandAll()

    def _create_top_item(self, group: GroupResult, version_names: List[str]) -> QTreeWidgetItem:
        """创建顶层行：编号 + ✓/✗ + 状态"""
        item = QTreeWidgetItem()

        # 列 0: 视频编号
        item.setText(0, group.file_id)
        item.setTextAlignment(0, Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)

        # 列 1~6: 各检测项 ✓/✗
        inconsistent_labels = []
        for c, (label, attr, formatter) in enumerate(COMPARE_COLUMNS, start=1):
            consistent = getattr(group, f"{attr}_consistent", True)
            if consistent:
                # 一致：✓ + tooltip 显示基准值
                first_vf = group.versions[0] if group.versions else None
                val_text = formatter(first_vf) if first_vf else "-"
                item.setText(c, "✓")
                item.setTextAlignment(c, Qt.AlignmentFlag.AlignCenter)
                item.setToolTip(c, f"所有版本一致：{val_text}")
                # 绿色文字
                item.setForeground(c, QColor("#137333"))
            else:
                # 不一致：✗ + tooltip 显示各版本值
                inconsistent_labels.append(label)
                vf_by_name = {v.version_name: v for v in group.versions}
                parts = []
                for vn in version_names:
                    vf = vf_by_name.get(vn)
                    parts.append(f"{vn}: {formatter(vf) if vf else '缺失'}")
                item.setText(c, "✗")
                item.setTextAlignment(c, Qt.AlignmentFlag.AlignCenter)
                item.setToolTip(c, "\n".join(parts))
                # 红色文字
                item.setForeground(c, QColor("#C5221F"))

        # 列 7: 状态
        if group.all_consistent:
            item.setText(COL_COUNT - 1, "✓ 一致")
            item.setForeground(COL_COUNT - 1, QColor("#137333"))
        else:
            n = len(inconsistent_labels)
            item.setText(COL_COUNT - 1, f"✗ 不一致 ({n} 项：{', '.join(inconsistent_labels)})")
            item.setForeground(COL_COUNT - 1, QColor("#C5221F"))
        item.setTextAlignment(COL_COUNT - 1, Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)

        # 整行 tooltip 显示简要说明
        item.setToolTip(0, "; ".join(group.messages))

        return item

    def _create_version_child(self, vf: VersionFile, group: GroupResult,
                               version_names: List[str]) -> QTreeWidgetItem:
        """创建子行：某版本的详细参数值"""
        child = QTreeWidgetItem()

        # 列 0: 版本名
        child.setText(0, vf.version_name)
        child.setTextAlignment(0, Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)

        # 列 1~6: 各项具体参数值
        for c, (label, attr, formatter) in enumerate(COMPARE_COLUMNS, start=1):
            val_text = formatter(vf)
            consistent = getattr(group, f"{attr}_consistent", True)
            child.setText(c, val_text)
            child.setTextAlignment(c, Qt.AlignmentFlag.AlignCenter)

            # 不一致项用红色高亮该版本的值
            if not consistent:
                child.setForeground(c, QColor("#C5221F"))

        # 列 7: 空（子行不需要状态列）
        child.setText(COL_COUNT - 1, "")

        # 整行 tooltip：完整文件路径
        if vf.error:
            child.setToolTip(0, f"文件：{vf.filepath}\n探测失败：{vf.error}")
            child.setForeground(0, QColor("#C5221F"))
        else:
            child.setToolTip(0, vf.filepath)

        return child

    # ── Excel 导出 ──────────────────────────────────────────
    def _export_excel(self):
        if not self._result:
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "导出对比结果", f"多版本对比_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel 文件 (*.xlsx)"
        )
        if not path:
            return
        try:
            self._write_excel(self._result, path)
            QMessageBox.information(self, "导出成功", f"已保存到：\n{path}")
        except Exception as e:
            logger.exception("导出 Excel 失败")
            QMessageBox.critical(self, "导出失败", f"导出 Excel 时出错：\n{e}")

    def _write_excel(self, result: CompareResult, path: str):
        """写入 Excel：每个视频一行（与界面表格一致），不一致项红色标注"""
        wb = Workbook()
        ws = wb.active
        ws.title = "多版本对比结果"

        # 样式
        title_font = Font(name="微软雅黑", size=16, bold=True, color="1A73E8")
        header_font = Font(name="微软雅黑", size=11, bold=True, color="1A1A2E")
        header_fill = PatternFill(start_color="F1F3F4", end_color="F1F3F4", fill_type="solid")
        pass_fill = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
        fail_fill = PatternFill(start_color="FCE8E6", end_color="FCE8E6", fill_type="solid")
        warn_font = Font(name="微软雅黑", size=10, color="C5221F", bold=True)
        ok_font = Font(name="微软雅黑", size=10, color="137333", bold=True)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        thin = Side(style="thin", color="DADCE0")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        headers = ["视频编号"] + [c[0] for c in COMPARE_COLUMNS] + ["状态", "说明"]
        n_cols = len(headers)

        # ── 标题 ──
        title = f"多版本视频对比报告 - {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
        title_cell = ws.cell(row=1, column=1, value=title)
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        # ── 摘要 ──
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
        summary = (
            f"共 {result.total_files} 个文件，"
            f"一致 {result.consistent_groups} 组，"
            f"不一致 {result.inconsistent_groups} 组"
        )
        sc = ws.cell(row=2, column=1, value=summary)
        sc.font = Font(name="微软雅黑", size=10, color="5F6368")
        sc.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 20

        # ── 表头 ──
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border
        ws.row_dimensions[3].height = 26
        ws.freeze_panes = "A4"

        # ── 数据行 ──
        sorted_groups = sorted(result.groups, key=lambda g: natural_sort_key(g.file_id))
        row_idx = 4
        for group in sorted_groups:
            self._write_excel_row(ws, row_idx, group, result.version_names,
                                  center, left, border, pass_fill, fail_fill, warn_font, ok_font)
            row_idx += 1

        # ── 列宽 ──
        widths = [22, 12, 10, 12, 14, 10, 14, 18, 50]
        for i, w in enumerate(widths[:n_cols], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        wb.save(path)

    def _write_excel_row(self, ws, row_idx, group, version_names, center, left, border,
                         pass_fill, fail_fill, warn_font, ok_font):
        """写入 Excel 单行：编号 + ✓/✗ + 状态 + 说明"""
        thin = Side(style="thin", color="DADCE0")
        cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # 视频编号
        c = ws.cell(row=row_idx, column=1, value=group.file_id)
        c.alignment = center
        c.border = cell_border

        # 各属性：一致显示基准值，不一致显示各版本值
        inconsistent_attrs = []
        vf_by_name = {v.version_name: v for v in group.versions}
        for i, (label, attr, formatter) in enumerate(COMPARE_COLUMNS):
            col = 2 + i
            consistent = getattr(group, f"{attr}_consistent", True)
            if consistent:
                first_vf = group.versions[0] if group.versions else None
                text = formatter(first_vf) if first_vf else "-"
            else:
                inconsistent_attrs.append(label)
                parts = []
                for vn in version_names:
                    vf = vf_by_name.get(vn)
                    parts.append(f"{vn}: {formatter(vf) if vf else '缺失'}")
                text = " / ".join(parts)
            cell = ws.cell(row=row_idx, column=col, value=text)
            cell.alignment = center
            cell.border = cell_border
            if not consistent:
                cell.font = warn_font
                cell.fill = fail_fill

        # 状态
        status_col = 2 + len(COMPARE_COLUMNS)
        status = "✓ 一致" if group.all_consistent else f"✗ 不一致 ({len(inconsistent_attrs)} 项)"
        sc = ws.cell(row=row_idx, column=status_col, value=status)
        sc.alignment = center
        sc.border = cell_border
        if group.all_consistent:
            sc.font = ok_font
            sc.fill = pass_fill
        else:
            sc.font = warn_font
            sc.fill = fail_fill

        # 说明
        detail_col = status_col + 1
        messages = group.messages if not group.all_consistent else ["所有版本完全一致"]
        dc = ws.cell(row=row_idx, column=detail_col, value="; ".join(messages))
        dc.alignment = left
        dc.border = cell_border
        if not group.all_consistent:
            dc.fill = fail_fill

    # ── 关闭 ────────────────────────────────────────────────
    def closeEvent(self, event):
        if self._thread and self._thread.isRunning():
            self._thread.wait(1000)
        super().closeEvent(event)
